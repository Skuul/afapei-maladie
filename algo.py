from collections import defaultdict, namedtuple
from datetime import datetime, timedelta
from typing import Dict, List
import openpyxl
from copy import copy
from os import sep as fs_sep

_Enregistrement = namedtuple('Enregistrement',
    ['Matricule',
    'Nom',
    'Prenom',
    'Code_service',
    'Service',
    'Date_planning',
    'Motif',
    'Libelle_du_motif',
    'Heures',
    'Arret_initial']
)
_Absence = namedtuple('Absence', ['debut', 'fin', 'motif'])

e = _Absence(0,0,0)

def _estJourSuivant(d1:datetime, d2:datetime):
    modified_date = copy(d1) + timedelta(days=1)
    s1 = datetime.strftime(modified_date, "%Y/%m/%d")
    s2 = datetime.strftime(d2, "%Y/%m/%d")
    return s1 == s2

def _filterAbsences(base: List[_Enregistrement]):
    d = defaultdict(list)
    for enr in base:
        d[enr.Matricule].append(enr)

    # Rend les données sous forme de générateur
    yield from map(
        lambda v: (v[0], sorted(v[1], key=lambda a: a.Date_planning)),
        d.items()
    )

def _getAbsences(absences: List[_Enregistrement]) -> List[_Absence]:
    absencesSalarie = []
    for absence in absences:
        if not len(absencesSalarie):
            absencesSalarie.append(
                _Absence(absence.Date_planning, absence.Date_planning, absence.Motif)
            )
            continue
        else:
            derniereAbsence = absencesSalarie.pop()
            estMemeMotif = derniereAbsence.motif == absence.Motif
            if estMemeMotif:
                jourSuivant = _estJourSuivant(derniereAbsence.fin, absence.Date_planning)
                if _estJourSuivant(derniereAbsence.fin, absence.Date_planning):
                    derniereAbsence = derniereAbsence._replace(fin=absence.Date_planning)
                    absencesSalarie.append(derniereAbsence)
            else:
                absencesSalarie.append(derniereAbsence)

                # Ajouter une absence
                absencesSalarie.append(
                    _Absence(absence.Date_planning, absence.Date_planning, absence.Motif)
                )
    return absencesSalarie

def _processBase(base: List[_Enregistrement]):
    absencesSalaries = dict()

    for matricule, absences in _filterAbsences(base):
        print(f'Looking at {matricule}...')
        absencesSalarie = _getAbsences(absences)
        absencesSalaries[matricule] = absencesSalarie

    return absencesSalaries

def algo(filename:str, outDir):
    wb = openpyxl.load_workbook(filename)
    sheetName = wb.sheetnames[0]
    sheet = wb[sheetName]
    
    base = []

    #process file
    it = iter(sheet.rows)
    # Ignore header
    next(it)
    for row in it:
        base.append(_Enregistrement(*(cell.value for cell in row)))

    absencesSalaries = _processBase(base)
    outputFileName = fs_sep.join([outDir, filename.split(fs_sep)[-1].split(".")[0]])
    _outputTxt(absencesSalaries, outputFileName+'.txt')
    _outputXlsx(absencesSalaries, outputFileName+'.xlsx')
    print(f'Done with {filename.split(fs_sep)[-1]}')
    # Write to file

def _outputTxt(absencesSalaries:Dict[int, _Absence], filename):
    with open(filename, 'w') as file:
        for matricule, absences in absencesSalaries.items():
            lines = (
                f'{matricule}: {_displayAbsence(absence)}\n'
                for absence in absences
            )
            file.writelines(lines)

def _displayAbsence(absence:_Absence):
    return absence.debut, absence.fin, absence.motif

def _outputXlsx(absencesSalaries:Dict[int, _Absence], filename):
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = 'Absences par salarié'
    
    # Headers
    ws.append((
        'Matricule',
        'Début',
        'Fin',
        'Motif'
    ))

    # Data
    for matricule, absences in absencesSalaries.items():
        for absence in absences:
            ws.append((matricule, *absence))
    
    wb.save(filename)